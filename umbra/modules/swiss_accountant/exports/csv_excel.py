"""
Export Manager for Swiss Accountant
Handles CSV and Excel exports of financial data for tax preparation and analysis.
"""
import csv
import json
from typing import Dict, List, Optional, Tuple, Any
from decimal import Decimal
from datetime import datetime, date
from io import StringIO, BytesIO
from pathlib import Path
import logging

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class ExportFormat:
    """Export format constants."""
    CSV = "csv"
    EXCEL = "xlsx"
    JSON = "json"


class ExportManager:
    """Manages data exports for Swiss Accountant."""
    
    def __init__(self):
        """Initialize export manager."""
        self.logger = logging.getLogger(__name__)
        
        # Export templates and configurations
        self.export_configs = {
            'tax_export': {
                'columns': [
                    'date', 'description', 'amount_chf', 'category', 'merchant',
                    'deduction_category', 'tax_deductible', 'deduction_amount',
                    'canton', 'notes', 'receipt_available'
                ],
                'headers': [
                    'Date', 'Description', 'Amount (CHF)', 'Expense Category', 'Merchant',
                    'Tax Category', 'Tax Deductible', 'Deductible Amount',
                    'Canton', 'Notes', 'Receipt Available'
                ]
            },
            'vat_export': {
                'columns': [
                    'date', 'merchant', 'amount_net', 'vat_rate', 'vat_amount', 
                    'amount_gross', 'business_percentage', 'deductible_vat',
                    'invoice_number', 'category'
                ],
                'headers': [
                    'Date', 'Merchant', 'Net Amount', 'VAT Rate (%)', 'VAT Amount',
                    'Gross Amount', 'Business %', 'Deductible VAT',
                    'Invoice Number', 'Category'
                ]
            },
            'expense_export': {
                'columns': [
                    'date', 'merchant', 'amount', 'currency', 'category',
                    'payment_method', 'account', 'notes', 'created_at'
                ],
                'headers': [
                    'Date', 'Merchant', 'Amount', 'Currency', 'Category',
                    'Payment Method', 'Account', 'Notes', 'Created At'
                ]
            },
            'reconciliation_export': {
                'columns': [
                    'expense_date', 'expense_amount', 'expense_merchant',
                    'transaction_date', 'transaction_amount', 'transaction_counterparty',
                    'match_type', 'confidence', 'match_status'
                ],
                'headers': [
                    'Expense Date', 'Expense Amount', 'Expense Merchant',
                    'Transaction Date', 'Transaction Amount', 'Transaction Counterparty',
                    'Match Type', 'Confidence', 'Status'
                ]
            }
        }
    
    def export_tax_data(self, 
                       db_manager,
                       user_id: str,
                       year: int,
                       format: str = ExportFormat.CSV,
                       canton: str = None) -> Dict[str, Any]:
        """Export tax-relevant data for the specified year.
        
        Args:
            db_manager: Database manager instance
            user_id: User identifier
            year: Tax year to export
            format: Export format (csv, xlsx, json)
            canton: Optional canton filter
            
        Returns:
            Dict with export result
        """
        try:
            # Get tax-relevant expenses
            query = """
                SELECT 
                    e.date_local,
                    e.merchant_text,
                    e.amount_cents,
                    e.currency,
                    e.category_code,
                    e.notes,
                    e.pro_pct,
                    e.payment_method,
                    m.canonical as merchant_canonical,
                    cm.deduction_category,
                    cm.confidence as mapping_confidence,
                    rd.file_name as receipt_file,
                    e.created_at
                FROM sa_expenses e
                LEFT JOIN sa_merchants m ON e.merchant_id = m.id
                LEFT JOIN sa_category_mappings cm ON e.category_code = cm.expense_category
                LEFT JOIN sa_documents d ON e.doc_id = d.id
                LEFT JOIN sa_receipts rd ON d.id = rd.doc_id
                WHERE e.user_id = ?
                AND strftime('%Y', e.date_local) = ?
                ORDER BY e.date_local, e.amount_cents DESC
            """
            
            params = [user_id, str(year)]
            if canton:
                query = query.replace("ORDER BY", "AND cm.canton = ? ORDER BY")
                params.insert(-1, canton)
            
            expenses = db_manager.query_all(query, params)
            
            if not expenses:
                return {
                    'success': False,
                    'error': f'No expenses found for year {year}',
                    'filename': None,
                    'data': None
                }
            
            # Process data for export
            export_data = self._prepare_tax_data(expenses)
            
            # Generate export based on format
            if format == ExportFormat.CSV:
                result = self._export_to_csv(export_data, 'tax_export')
            elif format == ExportFormat.EXCEL:
                result = self._export_to_excel(export_data, 'tax_export', f'Tax Export {year}')
            elif format == ExportFormat.JSON:
                result = self._export_to_json(export_data)
            else:
                return {
                    'success': False,
                    'error': f'Unsupported format: {format}'
                }
            
            if result['success']:
                filename = f'tax_export_{year}_{canton or "all"}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.{format}'
                result['filename'] = filename
                result['year'] = year
                result['canton'] = canton
                result['record_count'] = len(export_data)
            
            return result
            
        except Exception as e:
            self.logger.error(f"Tax data export failed: {e}")
            return {
                'success': False,
                'error': str(e),
                'filename': None,
                'data': None
            }
    
    def export_vat_data(self,
                       db_manager,
                       user_id: str,
                       period_start: date,
                       period_end: date,
                       format: str = ExportFormat.CSV) -> Dict[str, Any]:
        """Export VAT data for the specified period.
        
        Args:
            db_manager: Database manager instance
            user_id: User identifier
            period_start: Start date of period
            period_end: End date of period
            format: Export format
            
        Returns:
            Dict with export result
        """
        try:
            # Get VAT-relevant expenses
            expenses = db_manager.query_all("""
                SELECT 
                    e.date_local,
                    e.merchant_text,
                    e.amount_cents,
                    e.currency,
                    e.category_code,
                    e.vat_breakdown_json,
                    e.pro_pct,
                    e.notes,
                    m.canonical as merchant_canonical
                FROM sa_expenses e
                LEFT JOIN sa_merchants m ON e.merchant_id = m.id
                WHERE e.user_id = ?
                AND e.date_local BETWEEN ? AND ?
                AND e.pro_pct > 0
                ORDER BY e.date_local
            """, (user_id, period_start, period_end))
            
            # Process VAT data
            vat_data = self._prepare_vat_data(expenses)
            
            # Generate export
            if format == ExportFormat.CSV:
                result = self._export_to_csv(vat_data, 'vat_export')
            elif format == ExportFormat.EXCEL:
                result = self._export_to_excel(vat_data, 'vat_export', f'VAT Export {period_start} to {period_end}')
            elif format == ExportFormat.JSON:
                result = self._export_to_json(vat_data)
            else:
                return {
                    'success': False,
                    'error': f'Unsupported format: {format}'
                }
            
            if result['success']:
                filename = f'vat_export_{period_start}_{period_end}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.{format}'
                result['filename'] = filename
                result['period_start'] = period_start.isoformat()
                result['period_end'] = period_end.isoformat()
                result['record_count'] = len(vat_data)
            
            return result
            
        except Exception as e:
            self.logger.error(f"VAT data export failed: {e}")
            return {
                'success': False,
                'error': str(e),
                'filename': None,
                'data': None
            }
    
    def export_reconciliation_data(self,
                                 db_manager,
                                 user_id: str,
                                 session_id: int = None,
                                 format: str = ExportFormat.CSV) -> Dict[str, Any]:
        """Export reconciliation data.
        
        Args:
            db_manager: Database manager instance
            user_id: User identifier
            session_id: Optional specific session ID
            format: Export format
            
        Returns:
            Dict with export result
        """
        try:
            # Build query based on parameters
            if session_id:
                # Export specific session
                reconciliation_data = db_manager.query_all("""
                    SELECT 
                        e.date_local as expense_date,
                        e.amount_cents as expense_amount,
                        e.merchant_text as expense_merchant,
                        COALESCE(t.value_date, t.booking_date) as transaction_date,
                        t.amount_cents as transaction_amount,
                        t.counterparty as transaction_counterparty,
                        rm.match_type,
                        rm.confidence_score,
                        CASE 
                            WHEN rm.user_confirmed THEN 'Confirmed'
                            WHEN rm.user_rejected THEN 'Rejected'
                            ELSE 'Pending'
                        END as match_status
                    FROM sa_reconciliation_matches rm
                    JOIN sa_expenses e ON rm.expense_id = e.id
                    JOIN sa_transactions t ON rm.transaction_id = t.id
                    WHERE e.user_id = ?
                    ORDER BY rm.confidence_score DESC
                """, (user_id,))
            else:
                # Export all reconciliation data
                reconciliation_data = db_manager.query_all("""
                    SELECT 
                        e.date_local as expense_date,
                        e.amount_cents as expense_amount,
                        e.merchant_text as expense_merchant,
                        COALESCE(t.value_date, t.booking_date) as transaction_date,
                        t.amount_cents as transaction_amount,
                        t.counterparty as transaction_counterparty,
                        rm.match_type,
                        rm.confidence_score,
                        CASE 
                            WHEN rm.user_confirmed THEN 'Confirmed'
                            WHEN rm.user_rejected THEN 'Rejected'
                            ELSE 'Pending'
                        END as match_status
                    FROM sa_reconciliation_matches rm
                    JOIN sa_expenses e ON rm.expense_id = e.id
                    JOIN sa_transactions t ON rm.transaction_id = t.id
                    WHERE e.user_id = ?
                    ORDER BY e.date_local DESC
                """, (user_id,))
            
            # Process data
            processed_data = self._prepare_reconciliation_data(reconciliation_data)
            
            # Generate export
            if format == ExportFormat.CSV:
                result = self._export_to_csv(processed_data, 'reconciliation_export')
            elif format == ExportFormat.EXCEL:
                result = self._export_to_excel(processed_data, 'reconciliation_export', 'Reconciliation Export')
            else:
                return {
                    'success': False,
                    'error': f'Unsupported format: {format}'
                }
            
            if result['success']:
                session_suffix = f'_session_{session_id}' if session_id else '_all'
                filename = f'reconciliation_export{session_suffix}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.{format}'
                result['filename'] = filename
                result['session_id'] = session_id
                result['record_count'] = len(processed_data)
            
            return result
            
        except Exception as e:
            self.logger.error(f"Reconciliation data export failed: {e}")
            return {
                'success': False,
                'error': str(e),
                'filename': None,
                'data': None
            }
    
    def _prepare_tax_data(self, expenses: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Prepare expenses data for tax export."""
        try:
            prepared_data = []
            
            for expense in expenses:
                amount_chf = float(Decimal(expense['amount_cents']) / 100)
                
                # Determine tax deductibility
                deduction_category = expense.get('deduction_category', 'non_deductible')
                is_tax_deductible = deduction_category != 'non_deductible'
                
                # Calculate deductible amount
                if is_tax_deductible and expense.get('pro_pct', 0) > 0:
                    deductible_amount = amount_chf * (expense['pro_pct'] / 100)
                elif is_tax_deductible:
                    deductible_amount = amount_chf
                else:
                    deductible_amount = 0.0
                
                prepared_data.append({
                    'date': expense['date_local'],
                    'description': expense['merchant_text'] or expense.get('merchant_canonical', ''),
                    'amount_chf': amount_chf,
                    'category': expense.get('category_code', ''),
                    'merchant': expense.get('merchant_canonical', expense.get('merchant_text', '')),
                    'deduction_category': deduction_category,
                    'tax_deductible': 'Yes' if is_tax_deductible else 'No',
                    'deduction_amount': round(deductible_amount, 2),
                    'canton': '',  # Would be filled from user profile
                    'notes': expense.get('notes', ''),
                    'receipt_available': 'Yes' if expense.get('receipt_file') else 'No'
                })
            
            return prepared_data
            
        except Exception as e:
            self.logger.error(f"Tax data preparation failed: {e}")
            return []
    
    def _prepare_vat_data(self, expenses: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Prepare expenses data for VAT export."""
        try:
            prepared_data = []
            
            for expense in expenses:
                amount_chf = float(Decimal(expense['amount_cents']) / 100)
                business_pct = expense.get('pro_pct', 0)
                
                # Parse VAT breakdown
                vat_breakdown = {}
                if expense.get('vat_breakdown_json'):
                    try:
                        vat_breakdown = json.loads(expense['vat_breakdown_json'])
                    except json.JSONDecodeError:
                        pass
                
                # Extract VAT information
                vat_rate = vat_breakdown.get('rate', 8.1)  # Default Swiss standard rate
                
                # Calculate VAT amounts (assuming amount includes VAT)
                gross_amount = amount_chf
                net_amount = gross_amount / (1 + vat_rate / 100)
                vat_amount = gross_amount - net_amount
                
                # Calculate business portion
                business_net = net_amount * (business_pct / 100)
                business_vat = vat_amount * (business_pct / 100)
                
                prepared_data.append({
                    'date': expense['date_local'],
                    'merchant': expense.get('merchant_canonical', expense.get('merchant_text', '')),
                    'amount_net': round(net_amount, 2),
                    'vat_rate': vat_rate,
                    'vat_amount': round(vat_amount, 2),
                    'amount_gross': round(gross_amount, 2),
                    'business_percentage': business_pct,
                    'deductible_vat': round(business_vat, 2),
                    'invoice_number': '',  # Would need to be extracted from documents
                    'category': expense.get('category_code', '')
                })
            
            return prepared_data
            
        except Exception as e:
            self.logger.error(f"VAT data preparation failed: {e}")
            return []
    
    def _prepare_reconciliation_data(self, reconciliation_data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Prepare reconciliation data for export."""
        try:
            prepared_data = []
            
            for item in reconciliation_data:
                expense_amount = float(Decimal(item['expense_amount']) / 100)
                transaction_amount = float(Decimal(item['transaction_amount']) / 100)
                
                prepared_data.append({
                    'expense_date': item['expense_date'],
                    'expense_amount': expense_amount,
                    'expense_merchant': item['expense_merchant'],
                    'transaction_date': item['transaction_date'],
                    'transaction_amount': transaction_amount,
                    'transaction_counterparty': item['transaction_counterparty'],
                    'match_type': item['match_type'].title(),
                    'confidence': round(item['confidence_score'], 3),
                    'match_status': item['match_status']
                })
            
            return prepared_data
            
        except Exception as e:
            self.logger.error(f"Reconciliation data preparation failed: {e}")
            return []
    
    def _export_to_csv(self, data: List[Dict[str, Any]], export_type: str) -> Dict[str, Any]:
        """Export data to CSV format."""
        try:
            if not data:
                return {
                    'success': False,
                    'error': 'No data to export'
                }
            
            config = self.export_configs.get(export_type, {})
            columns = config.get('columns', list(data[0].keys()))
            headers = config.get('headers', columns)
            
            # Create CSV content
            output = StringIO()
            writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_MINIMAL)
            
            # Write header
            writer.writerow(headers)
            
            # Write data
            for row in data:
                csv_row = []
                for col in columns:
                    value = row.get(col, '')
                    # Handle None values and convert to string
                    if value is None:
                        value = ''
                    elif isinstance(value, (int, float, Decimal)):
                        value = str(value)
                    csv_row.append(value)
                writer.writerow(csv_row)
            
            csv_content = output.getvalue()
            output.close()
            
            return {
                'success': True,
                'content': csv_content,
                'format': 'csv',
                'size_bytes': len(csv_content.encode('utf-8'))
            }
            
        except Exception as e:
            self.logger.error(f"CSV export failed: {e}")
            return {
                'success': False,
                'error': str(e)
            }
    
    def _export_to_excel(self, 
                        data: List[Dict[str, Any]], 
                        export_type: str, 
                        sheet_name: str) -> Dict[str, Any]:
        """Export data to Excel format."""
        try:
            if not EXCEL_AVAILABLE:
                return {
                    'success': False,
                    'error': 'Excel export not available (openpyxl not installed)'
                }
            
            if not data:
                return {
                    'success': False,
                    'error': 'No data to export'
                }
            
            config = self.export_configs.get(export_type, {})
            columns = config.get('columns', list(data[0].keys()))
            headers = config.get('headers', columns)
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # Style definitions
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Write headers
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Write data
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, col_name in enumerate(columns, 1):
                    value = row_data.get(col_name, '')
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    
                    # Auto-format numbers
                    if isinstance(value, (int, float, Decimal)) and col_name.endswith(('amount', 'vat', 'total')):
                        cell.number_format = '#,##0.00'
                    elif col_name.endswith('date'):
                        cell.number_format = 'DD.MM.YYYY'
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            excel_content = output.getvalue()
            output.close()
            
            return {
                'success': True,
                'content': excel_content,
                'format': 'xlsx',
                'size_bytes': len(excel_content)
            }
            
        except Exception as e:
            self.logger.error(f"Excel export failed: {e}")
            return {
                'success': False,
                'error': str(e)
            }
    
    def _export_to_json(self, data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Export data to JSON format."""
        try:
            json_content = json.dumps({
                'export_timestamp': datetime.now().isoformat(),
                'record_count': len(data),
                'data': data
            }, indent=2, default=str)
            
            return {
                'success': True,
                'content': json_content,
                'format': 'json',
                'size_bytes': len(json_content.encode('utf-8'))
            }
            
        except Exception as e:
            self.logger.error(f"JSON export failed: {e}")
            return {
                'success': False,
                'error': str(e)
            }
    
    def get_export_summary(self, db_manager, user_id: str) -> Dict[str, Any]:
        """Get summary of available data for export.
        
        Args:
            db_manager: Database manager instance
            user_id: User identifier
            
        Returns:
            Dict with export summary
        """
        try:
            # Get expenses summary
            expenses_summary = db_manager.query_one("""
                SELECT 
                    COUNT(*) as total_expenses,
                    MIN(date_local) as earliest_date,
                    MAX(date_local) as latest_date,
                    SUM(amount_cents) as total_amount_cents,
                    COUNT(DISTINCT strftime('%Y', date_local)) as years_available
                FROM sa_expenses
                WHERE user_id = ?
            """, (user_id,))
            
            # Get reconciliation summary
            reconciliation_summary = db_manager.query_one("""
                SELECT 
                    COUNT(*) as total_matches,
                    SUM(CASE WHEN user_confirmed = TRUE THEN 1 ELSE 0 END) as confirmed_matches
                FROM sa_reconciliation_matches rm
                JOIN sa_expenses e ON rm.expense_id = e.id
                WHERE e.user_id = ?
            """, (user_id,))
            
            # Get available years
            available_years = db_manager.query_all("""
                SELECT DISTINCT strftime('%Y', date_local) as year, COUNT(*) as expense_count
                FROM sa_expenses
                WHERE user_id = ?
                GROUP BY strftime('%Y', date_local)
                ORDER BY year DESC
            """, (user_id,))
            
            return {
                'total_expenses': expenses_summary['total_expenses'] or 0,
                'earliest_date': expenses_summary['earliest_date'],
                'latest_date': expenses_summary['latest_date'],
                'total_amount_chf': float(Decimal(expenses_summary['total_amount_cents'] or 0) / 100),
                'years_available': [
                    {
                        'year': int(year['year']),
                        'expense_count': year['expense_count']
                    }
                    for year in available_years
                ],
                'reconciliation': {
                    'total_matches': reconciliation_summary['total_matches'] or 0,
                    'confirmed_matches': reconciliation_summary['confirmed_matches'] or 0
                },
                'supported_formats': ['csv', 'xlsx', 'json'],
                'export_types': list(self.export_configs.keys())
            }
            
        except Exception as e:
            self.logger.error(f"Export summary failed: {e}")
            return {
                'total_expenses': 0,
                'earliest_date': None,
                'latest_date': None,
                'total_amount_chf': 0.0,
                'years_available': [],
                'reconciliation': {'total_matches': 0, 'confirmed_matches': 0},
                'supported_formats': ['csv', 'xlsx', 'json'],
                'export_types': list(self.export_configs.keys()),
                'error': str(e)
            }


# Factory function for easy import
def create_export_manager() -> ExportManager:
    """Create export manager instance."""
    return ExportManager()
